'''
Create XII style test index file for DQC conf tests

Procedure:

from https://xbrl.us/data-quality/certification/process/

1) obtain accession numbers and edit into SQL query of arelle database (to get filing date, name and entry URL):

   select accession_number, form_type, filing_date, entity_name, entry_url, namespace from filing_edgar fe
   join report r on r.filing_fk = fe.filing_pk join report_edgar re on re.report_pk = r.report_pk
   join document d  on re.standard_schema_doc_fk = d.document_pk
   where accession_number in (  ... list of accession numbers )
   
   save into file ACCESSION_INSTANCES
   
2) using file ACCESSION_INSTANCES (produced from simple regex editing of the certification results file)

   run this program to produce the test case file.
   
3) to generate an annual index:

   select accession_number, form_type, filing_date, entity_name, entry_url, namespace from filing_edgar fe
   join report r on r.filing_fk = fe.filing_pk join report_edgar re on re.report_pk = r.report_pk
   join document d  on re.standard_schema_doc_fk = d.document_pk
   where filing_date >= '2020-01-01' and filing_date < '2020-12-31' and namespace = 'http://fasb.org/us-gaap/2020-01-31'


'''
import os, csv, re
from openpyxl import load_workbook
from collections import OrderedDict

EXCELWB = "/Users/hermf/Documents/projects/SEC/25.1/DQCRT/DQCRuleManifest-2025.xlsx"
TESTCASE = "/Users/hermf/Documents/projects/SEC/25.1/DQCRT/dqc-testcases/DQC-certification-test.xml"
TESTCASE_SPLIT_CPUs = 1
TESTCASE = "/Users/hermf/Documents/projects/SEC/25.1/DQCRT/dqc-testcases/DQCRT-{}-test-{}-{}.xml"
TESTCASE_SPLIT_CPUs = 8

ENTRY_URL_PATTERN = re.compile(r"https?://www.sec.gov/Archives/edgar/data/[0-9]+/([0-9]{10})(2[34])([0-9]{6})/.*")

# MATCHING_ERROR_CODES = r"DQCRT.US.(0001|0004|0005|0006|0008|0009|0013|0014|0015|0033|0036|0041|0043|0044|0045|0046|0047|0048|0051|0052|0053|0054|0055|0057||0060||0061|0062|0065|0068|0069|0070|0071|0072|0073||0076|0077|0078|0079||0084|0085|0089|0090|0091|0095|0098|0099|0108|0109|0112|0118|0119|0123|0126|0128|0133|0134|0135|0136|0137|0141).*"
MATCHING_ERROR_CODES = r"DQC.US.(0001|0004|0005|0006|0008|0009|0013|0014|0015|0033|0036|0041|0043|0044|0045|0046|0047|0048|0051|0052|0053|0054|0055|0057||0060||0061|0062|0065|0068|0069|0070|0071|0072|0073||0076|0077|0078|0079||0084|0085|0089|0090|0091|0095|0098|0099|0108|0109|0112|0118|0119|0123|0126|0128|0133|0134|0135|0136|0137|0141).*"

dqcRuleFilter = {"more": "XULE:2023|.*", "less": "XULE:2023|.*", "python": "XULE:9999|.*"}

testcaseVariation = 1
wb = load_workbook(EXCELWB, data_only=True) 
ws = wb.active
# more runs all the 0015 variations, less runs only one 0015 variation
for yr, urlColLtr in ((2024, 'D'), (2025, 'H')):
    for testLen in ("more", "less", "python"):
        ruleId = 0
        numAssessions = 0
        outFile = 0
        fw = None
        variationInfo = OrderedDict()
        num0015variations = 99999 if testLen == "more" else 1
        for rowNum in range(2, ws.max_row + 1):
            entryURL = ws[f"{urlColLtr}{rowNum}"].value
            rule_code = ws[f"A{rowNum}"].value # .replace("DQC.US","DQCRT.US")
            if entryURL and not rule_code.startswith("#"):  # is a commented out row
                if ENTRY_URL_PATTERN.match(entryURL):
                    match = ENTRY_URL_PATTERN.match(entryURL)
                    accession = f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
                    zipURL = entryURL.rpartition("/")[0] + "/" + accession + "-xbrl.zip" 
                else:
                    accession = f"test{rowNum-1}"
                    zipURL = entryURL
                if not entryURL.startswith("http"):
                    zipURL = entryURL = os.path.join(f"conf-{yr}", entryURL)
                if rule_code == "DQC.US.0015" and testLen != "more": # "DQCRT.US.0015" and testLen == "less":
                    num0015variations -= 1
                    if num0015variations < 0:
                        continue # skip more 0015 for shorter test suite
                error_code = rule_code + "." + str(ws[f"B{rowNum}"].value)
                count = 1
                numAssessions += 1
                variationInfo[rowNum] = {
                    "accession": accession,
                    "name": entryURL.rpartition("/")[2],
                    "url": entryURL,
                    "zipURL": zipURL,
                    "errors": []}
                variationInfo[rowNum]["errors"].append((error_code, int(count)))
                if numAssessions:
                    splitFileAt = int((numAssessions + TESTCASE_SPLIT_CPUs - 1) / TESTCASE_SPLIT_CPUs)
            
        for i, row in enumerate(variationInfo.values()):
            if i % splitFileAt == 0:
                if fw is not None:
                    fw.write(
    """</testcase>
    """)
                    fw.close()
                    fw = None
                outFile += 1
                fw = open(TESTCASE.format(yr, testLen, outFile), "w")
                fw.write(
"""<?xml version="1.0" encoding="UTF-8"?>
<?xml-stylesheet type="text/xsl" href="../../infrastructure/test.xsl"?>
<testcase
    xmlns="http://xbrl.org/2008/conformance">
  <creator>
    <name>EDGAR Team based on DQC certification published by XBRL-US:  https://xbrl.us/data-quality/certification/process/</name>
    <email>StructuredData@sec.gov</email>
  </creator>
  <number>DQCRT</number>
  <name>XBRL_US DQC Certification Tests</name>
  <description>
    DQCRT rules for 2025."""
#"""
#    Adaptation of DQC certification test file to XII standard test case file."""
"""
    All us-gaap/2023-2024 filings"""
"""
    Results element @blockedMessageCodes filters unrelated test results.
    Results error elements are intended to match all detected errors.
    Counts of multiply-occurring errors are not expected to match.
    
    Note that the parameter xml elements below have their default prefix not assigned to a namespace, so that the parameter name 
    is not given a QName with namespace (of the conformance suite) and instead parameter name has a no-namespace QName value.
  </description>
""")
            fw.write(
"""  <variation id="v-{0}">
    <description>{1}, {2}, {3}, {4}</description>
    <data>
      <parameter name="dqcRuleFilter" value="{6}" />
      <instance readMeFirst="true">{4}</instance>
    </data>
    <result blockedMessageCodes="(?!{5})">
"""         .format(testcaseVariation, row["name"].replace("&","&amp;"), row["accession"], row["url"], row["zipURL"], MATCHING_ERROR_CODES,
                    dqcRuleFilter[testLen]))
            testcaseVariation += 1
            for error_code, count in sorted(row["errors"], key=lambda i:i[0]):
                for j in range(count):
                    fw.write(
"""      <error>{}</error>
"""                     .format(error_code))
            fw.write(
"""    </result>
  </variation>
""")
        if fw is not None:
            fw.write(
"""</testcase>
""")
        fw.close()
    print(f"Finished \"{testLen}\" testcase, {numAssessions} variations")